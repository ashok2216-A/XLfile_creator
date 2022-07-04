
import openpyxl as opxl

workb = opxl.Workbook()


activeSheet = workb.active


cell1 = activeSheet.cell(row=1, column=1)
cell2 = activeSheet.cell(row=2, column=1)
cell3 = activeSheet.cell(row=3, column=1)
cell4 = activeSheet.cell(row=4, column=1)
cell5 = activeSheet.cell(row=5, column=1)
cell6 = activeSheet.cell(row=6, column=1)

cell1.value = "Name"
cell2.value = "Hi"
cell3.value = "Apple"
cell4.value = "Orange"
cell5.value = "Good"
cell6.value = "1000"

cell3 = activeSheet['C1']
cell3.value = "1000"

cell3 = activeSheet['C2']
cell3.value = "565656"


workb.save("file1.xlsx")